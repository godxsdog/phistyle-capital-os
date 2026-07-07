from decimal import Decimal

import pytest


pytest.importorskip("openpyxl")
pytest.importorskip("sqlalchemy")

from openpyxl import Workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from backend.app.commands.import_rates_excel import import_rates_excel
from shared.database.base import Base
from shared.models import point_wallet  # noqa: F401
from shared.models.point_wallet import PointProgram, PurchaseOffer, TransferRule


def make_session():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
        future=True,
    )
    Base.metadata.create_all(engine)
    session_factory = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    return session_factory()


def write_workbook(path):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "換算表"
    ws.cell(row=1, column=5).value = "全球航司"
    ws.cell(row=1, column=6).value = "比例"
    ws.cell(row=1, column=7).value = "加贈"
    ws.cell(row=2, column=5).value = "Qatar"
    ws.cell(row=2, column=6).value = 70
    ws.cell(row=2, column=7).value = 1
    ws.cell(row=3, column=11).value = "AS官方特賣"
    ws.cell(row=4, column=11).value = 0.7
    ws.cell(row=4, column=12).value = "0.54/哩"
    ws.cell(row=5, column=11).value = "60000萬豪=25000AS"
    ws.cell(row=6, column=12).value = "https://example.test/rates"
    workbook.create_sheet("QR").cell(row=1, column=1).value = "自由文字"
    workbook.save(path)


def seed_programs(session):
    for name in ("平安萬里通", "Qatar", "萬豪", "AS"):
        session.add(PointProgram(name=name, kind="other"))
    session.commit()


def test_rates_excel_import_dry_run_commit_and_idempotency(tmp_path):
    path = tmp_path / "rates.xlsx"
    write_workbook(path)
    session = make_session()
    seed_programs(session)

    dry_run = import_rates_excel(session, path, commit=False)
    assert len(dry_run["will_create"]["transfer_rules"]) == 2
    assert len(dry_run["will_create"]["purchase_offers"]) == 1
    assert dry_run["will_create"]["transfer_rules"][1]["rule_kind"] == "threshold_block"
    assert dry_run["will_create"]["transfer_rules"][1]["block_bonus_points"] == "5000"

    first = import_rates_excel(session, path, commit=True)
    second = import_rates_excel(session, path, commit=True)

    assert first["created"] == {"transfer_rules": 2, "purchase_offers": 1}
    assert second["created"] == {"transfer_rules": 0, "purchase_offers": 0}
    assert len(session.scalars(select(TransferRule)).all()) == 2
    assert len(session.scalars(select(PurchaseOffer)).all()) == 1
    threshold_rule = session.scalar(select(TransferRule).where(TransferRule.rule_kind == "threshold_block"))
    assert threshold_rule.block_size == Decimal("60000.00")
    assert threshold_rule.block_bonus_points == Decimal("5000.00")
