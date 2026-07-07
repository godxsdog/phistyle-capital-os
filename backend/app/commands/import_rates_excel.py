from __future__ import annotations

import argparse
from dataclasses import asdict, dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
import json
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session, sessionmaker

from shared.database.base import Base
from shared.database.session import SessionLocal
from shared.models.point_wallet import PointProgram, PurchaseOffer, TransferRule
from shared.services.point_wallet_service import create_purchase_offer, create_transfer_rule


NOTE_PREFIX = "excel匯入-未查證"
WANLITONG_ALIASES = ("平安萬里通", "Wanlitong", "萬里通")
PROGRAM_ALIASES: dict[str, tuple[str, ...]] = {
    "ALL 雅高": ("ALL 雅高", "Accor ALL", "ALL Accor"),
    "萬豪": ("萬豪", "Marriott Bonvoy", "Marriott"),
    "AS": ("AS", "Alaska Airlines", "Alaska Mileage Plan"),
    "亞萬": ("亞萬", "Asia Miles", "Cathay Asia Miles"),
    "日航": ("日航", "JAL", "Japan Airlines"),
    "芬蘭": ("芬蘭", "Finnair"),
    "英航": ("英航", "BA Avios", "British Airways"),
    "Qatar": ("Qatar", "Qatar Avios", "QR"),
    "哥倫LM": ("哥倫LM", "LM", "LifeMiles", "Avianca LifeMiles"),
    "加拿大AC": ("加拿大AC", "AC", "Aeroplan"),
    "新加坡": ("新加坡", "KrisFlyer", "Singapore Airlines"),
    "美聯航": ("美聯航", "United", "United MileagePlus", "UA"),
    "長榮": ("長榮", "EVA Air", "Infinity MileageLands"),
    "FB": ("FB", "Flying Blue"),
    "澳航": ("澳航", "Qantas"),
    "土航": ("土航", "Turkish Airlines", "Miles&Smiles"),
    "阿提哈德": ("阿提哈德", "Etihad"),
    "阿聯求": ("阿聯求", "阿聯酋", "Emirates"),
    "UA": ("UA", "United", "United MileagePlus", "美聯航"),
}


@dataclass(frozen=True)
class TransferRuleCandidate:
    from_program_name: str
    to_program_name: str
    ratio_from: str
    ratio_to: str
    bonus_pct: str
    min_transfer: str | None
    valid_from: str
    valid_until: str | None
    rule_kind: str
    block_size: str | None
    block_bonus_points: str | None
    transfer_days_note: str
    source_url: str | None
    sheet: str
    row: int


@dataclass(frozen=True)
class PurchaseOfferCandidate:
    program_name: str
    kind: str
    base_price: str
    currency: str
    bonus_pct: str
    start_date: str
    end_date: str | None
    source_note: str
    source_url: str | None
    sheet: str
    row: int


def main() -> None:
    parser = argparse.ArgumentParser(description="Dry-run or import manually maintained Point Wallet rates from Excel.")
    parser.add_argument("--file", default="data-rescue/wanlitong_rates.xlsx", help="Path to wanlitong_rates.xlsx.")
    parser.add_argument("--commit", action="store_true", help="Persist clear candidates. Default is dry-run only.")
    parser.add_argument("--database-url", help="Optional database URL override for local verification.")
    args = parser.parse_args()

    session_factory = _session_factory(args.database_url)
    session = session_factory()
    try:
        result = import_rates_excel(session, Path(args.file), commit=args.commit)
        print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))
    finally:
        session.close()


def import_rates_excel(session: Session, path: Path, *, commit: bool = False) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True)
    candidates, warnings = extract_candidates(workbook)
    programs = list(session.scalars(select(PointProgram)))
    transfer_to_create, purchase_to_create, db_warnings = resolve_against_programs(candidates, programs)
    warnings.extend(db_warnings)
    created = {"transfer_rules": 0, "purchase_offers": 0}
    existing: list[dict[str, Any]] = []
    if commit:
        for candidate, from_program, to_program in transfer_to_create:
            if _transfer_exists(session, candidate, from_program.id, to_program.id):
                existing.append({"kind": "transfer_rule", "candidate": asdict(candidate)})
                continue
            create_transfer_rule(
                session,
                from_program_id=from_program.id,
                to_program_id=to_program.id,
                ratio_from=Decimal(candidate.ratio_from),
                ratio_to=Decimal(candidate.ratio_to),
                bonus_pct=Decimal(candidate.bonus_pct),
                min_transfer=Decimal(candidate.min_transfer) if candidate.min_transfer else None,
                transfer_days_note=candidate.transfer_days_note,
                valid_from=date.fromisoformat(candidate.valid_from),
                valid_until=date.fromisoformat(candidate.valid_until) if candidate.valid_until else None,
                rule_kind=candidate.rule_kind,
                block_size=Decimal(candidate.block_size) if candidate.block_size else None,
                block_bonus_points=Decimal(candidate.block_bonus_points) if candidate.block_bonus_points else None,
                source_url=candidate.source_url,
            )
            created["transfer_rules"] += 1
        for candidate, program in purchase_to_create:
            if _offer_exists(session, candidate, program.id):
                existing.append({"kind": "purchase_offer", "candidate": asdict(candidate)})
                continue
            create_purchase_offer(
                session,
                program_id=program.id,
                kind=candidate.kind,
                base_price=Decimal(candidate.base_price),
                currency=candidate.currency,
                bonus_pct=Decimal(candidate.bonus_pct),
                start_date=date.fromisoformat(candidate.start_date),
                end_date=date.fromisoformat(candidate.end_date) if candidate.end_date else None,
                source_note=candidate.source_note,
                source_url=candidate.source_url,
            )
            created["purchase_offers"] += 1
    return {
        "mode": "commit" if commit else "dry-run",
        "will_create": {
            "transfer_rules": [asdict(candidate) for candidate, _, _ in transfer_to_create],
            "purchase_offers": [asdict(candidate) for candidate, _ in purchase_to_create],
        },
        "existing_skipped": existing,
        "created": created,
        "warnings": warnings,
    }


def extract_candidates(workbook) -> tuple[dict[str, list[Any]], list[str]]:
    warnings: list[str] = []
    transfer_rules: list[TransferRuleCandidate] = []
    purchase_offers: list[PurchaseOfferCandidate] = []
    source_url = _first_url(workbook)
    sheet = workbook["換算表"] if "換算表" in workbook.sheetnames else None
    if sheet is None:
        return {"transfer_rules": [], "purchase_offers": []}, ["缺少「換算表」分頁，未萃取。"]

    for row_index in range(2, sheet.max_row + 1):
        target = _clean(sheet.cell(row=row_index, column=5).value)
        ratio = _decimal(sheet.cell(row=row_index, column=6).value)
        multiplier = _decimal(sheet.cell(row=row_index, column=7).value)
        if target and ratio is not None and multiplier is not None and ratio > 0 and multiplier > 0:
            transfer_rules.append(
                TransferRuleCandidate(
                    from_program_name="平安萬里通",
                    to_program_name=target,
                    ratio_from=_decimal_text(ratio),
                    ratio_to="1",
                    bonus_pct=_decimal_text((multiplier - Decimal("1")) * Decimal("100")),
                    min_transfer=None,
                    valid_from="2026-07-07",
                    valid_until=None,
                    rule_kind="linear",
                    block_size=None,
                    block_bonus_points=None,
                    transfer_days_note=f"{NOTE_PREFIX}; 換算表 row {row_index}",
                    source_url=source_url,
                    sheet=sheet.title,
                    row=row_index,
                )
            )
        elif any(sheet.cell(row=row_index, column=col).value is not None for col in (5, 6, 7)):
            warnings.append(f"換算表 row {row_index}: 全球航司欄位不完整或非數字，已跳過。")

    purchase_section = False
    for row_index in range(1, sheet.max_row + 1):
        for col_index in range(1, sheet.max_column + 1):
            value = _clean(sheet.cell(row=row_index, column=col_index).value)
            if "AS官方特賣" in value:
                purchase_section = True
                continue
            if purchase_section:
                price_match = re.fullmatch(r"([0-9]+(?:\.[0-9]+)?)/哩", value)
                if price_match:
                    tier = _clean(sheet.cell(row=row_index, column=max(1, col_index - 1)).value)
                    purchase_offers.append(
                        PurchaseOfferCandidate(
                            program_name="AS",
                            kind="promo",
                            base_price=_decimal_text(Decimal(price_match.group(1))),
                            currency="TWD",
                            bonus_pct="0",
                            start_date="2026-07-07",
                            end_date=None,
                            source_note=f"{NOTE_PREFIX}; AS官方特賣 階梯={tier or '未標示'}; 換算表 row {row_index}",
                            source_url=None,
                            sheet=sheet.title,
                            row=row_index,
                        )
                    )

    for ws in workbook.worksheets:
        if ws.title == "換算表":
            continue
        warnings.append(f"{ws.title}: 自由文字/備忘分頁，未萃取結構化規則。")
    for ws in workbook.worksheets:
        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for value in row:
                text = _clean(value)
                match = re.search(r"(\d{4,})萬豪[=＝](\d{4,})(AS|UA)", text, flags=re.IGNORECASE)
                if not match:
                    continue
                block_size = Decimal(match.group(1))
                total_received = Decimal(match.group(2))
                to_name = match.group(3).upper()
                base_received = (block_size / Decimal("3")).quantize(Decimal("1"))
                block_bonus = total_received - base_received
                if block_bonus <= 0:
                    warnings.append(f"{ws.title} row {row_index}: 萬豪門檻文字無法形成正數加贈，已跳過：{text}")
                    continue
                transfer_rules.append(
                    TransferRuleCandidate(
                        from_program_name="萬豪",
                        to_program_name=to_name,
                        ratio_from="3",
                        ratio_to="1",
                        bonus_pct="0",
                        min_transfer=None,
                        valid_from="2026-07-07",
                        valid_until=None,
                        rule_kind="threshold_block",
                        block_size=_decimal_text(block_size),
                        block_bonus_points=_decimal_text(block_bonus),
                        transfer_days_note=f"{NOTE_PREFIX}; {text}; {ws.title} row {row_index}",
                        source_url=source_url,
                        sheet=ws.title,
                        row=row_index,
                    )
                )
    return {"transfer_rules": transfer_rules, "purchase_offers": purchase_offers}, warnings


def resolve_against_programs(candidates: dict[str, list[Any]], programs: list[PointProgram]) -> tuple[list[tuple[TransferRuleCandidate, PointProgram, PointProgram]], list[tuple[PurchaseOfferCandidate, PointProgram]], list[str]]:
    warnings: list[str] = []
    transfer_rules: list[tuple[TransferRuleCandidate, PointProgram, PointProgram]] = []
    purchase_offers: list[tuple[PurchaseOfferCandidate, PointProgram]] = []
    for candidate in candidates["transfer_rules"]:
        from_program = _find_program(programs, candidate.from_program_name, WANLITONG_ALIASES)
        to_program = _find_program(programs, candidate.to_program_name, PROGRAM_ALIASES.get(candidate.to_program_name, (candidate.to_program_name,)))
        if from_program is None or to_program is None:
            warnings.append(
                f"{candidate.sheet} row {candidate.row}: 找不到既有 program，已跳過 transfer {candidate.from_program_name} → {candidate.to_program_name}。"
            )
            continue
        transfer_rules.append((candidate, from_program, to_program))
    for candidate in candidates["purchase_offers"]:
        program = _find_program(programs, candidate.program_name, PROGRAM_ALIASES.get(candidate.program_name, (candidate.program_name,)))
        if program is None:
            warnings.append(f"{candidate.sheet} row {candidate.row}: 找不到既有 program，已跳過 offer {candidate.program_name}。")
            continue
        purchase_offers.append((candidate, program))
    return transfer_rules, purchase_offers, warnings


def _transfer_exists(session: Session, candidate: TransferRuleCandidate, from_program_id: int, to_program_id: int) -> bool:
    return (
        session.scalar(
            select(TransferRule).where(
                TransferRule.from_program_id == from_program_id,
                TransferRule.to_program_id == to_program_id,
                TransferRule.ratio_from == Decimal(candidate.ratio_from),
                TransferRule.ratio_to == Decimal(candidate.ratio_to),
                TransferRule.bonus_pct == Decimal(candidate.bonus_pct),
                TransferRule.rule_kind == candidate.rule_kind,
                TransferRule.transfer_days_note == candidate.transfer_days_note,
            )
        )
        is not None
    )


def _offer_exists(session: Session, candidate: PurchaseOfferCandidate, program_id: int) -> bool:
    return (
        session.scalar(
            select(PurchaseOffer).where(
                PurchaseOffer.program_id == program_id,
                PurchaseOffer.kind == candidate.kind,
                PurchaseOffer.base_price == Decimal(candidate.base_price),
                PurchaseOffer.currency == candidate.currency,
                PurchaseOffer.source_note == candidate.source_note,
            )
        )
        is not None
    )


def _find_program(programs: list[PointProgram], label: str, aliases: tuple[str, ...]) -> PointProgram | None:
    alias_keys = {_key(alias) for alias in aliases}
    alias_keys.add(_key(label))
    for program in programs:
        if _key(program.name) in alias_keys:
            return program
    return None


def _first_url(workbook) -> str | None:
    for ws in workbook.worksheets:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                text = _clean(value)
                if text.startswith("http://") or text.startswith("https://"):
                    return text
    return None


def _decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _decimal_text(value: Decimal) -> str:
    normalized = value.quantize(Decimal("0.000001")).normalize()
    return format(normalized, "f")


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _key(value: str) -> str:
    return re.sub(r"\s+", "", value).lower()


def _session_factory(database_url: str | None):
    if not database_url:
        return SessionLocal
    engine = create_engine(database_url, future=True)
    if database_url.startswith("sqlite"):
        Base.metadata.create_all(engine)
    return sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)


if __name__ == "__main__":
    main()
